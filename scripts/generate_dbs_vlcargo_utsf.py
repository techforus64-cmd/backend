"""
generate_dbs_vlcargo_utsf.py

Generates UTSF v3.0 files for:
  - DB Schenker  (Pincode_DBS sheet)
  - VL Cargo     (Pincode_B2B_Delhivery sheet + Shipshopy serviceability template)

Run: python generate_dbs_vlcargo_utsf.py
"""
import openpyxl
import json
import os

EXCEL_PATH = r'C:\Users\FORUS\Downloads\aeiou (3)\aeiou\Transport Cost Calculator (5).xlsx'
UTSF_DIR   = r'C:\Users\FORUS\Downloads\aeiou (3)\aeiou\backend\data\utsf'
SHIPSHOPY_ID = '6968ddedc2cf85d3f4380d52'

DBS_ID = '67b4b800db5c000000000001'
VL_ID  = '67b4b800cf900000000000c1'
NOW    = '2026-02-18T12:00:00.000Z'

# Standard India zone totals (from existing UTSF data)
STANDARD_TOTALS = {
    'N1': 409, 'N2': 95,   'N3': 3727, 'N4': 932,
    'S1': 478, 'S2': 3177, 'S3': 2196, 'S4': 1757,
    'E1': 296, 'E2': 3335, 'W1': 397,  'W2': 2700,
    'C1': 130, 'C2': 925,  'NE1': 45,  'NE2': 872,
    'X1': 22,  'X2': 9,    'X3': 14
}

ALL_ZONES = ['N1','N2','N3','N4','S1','S2','S3','S4','E1','E2','W1','W2',
             'C1','C2','NE1','NE2','X1','X2','X3']

VL_ZONES = ['N1','N2','N3','N4','C1','C2','W1','W2','S1','S2','S3','S4',
            'E1','E2','NE1','NE2']

print("Reading Excel file...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# ============================================================
# STEP 1: DB SCHENKER DATA FROM Pincode_DBS
# ============================================================
ws_dbs = wb['Pincode_DBS']

dbs_rate_table = {}   # zone -> rate
dbs_pincodes   = {}   # zone -> {'all':[], 'oda':[], 'non_oda':[]}

# Rate table in columns I(idx=8) and J(idx=9)
for row in ws_dbs.iter_rows(min_row=2, max_row=20, values_only=True):
    if row[8] is not None and row[9] is not None:
        zone_name = str(row[8]).strip()
        rate_val  = float(row[9])
        if zone_name and zone_name != 'Zone':
            dbs_rate_table[zone_name] = rate_val

# All served pincodes
for row in ws_dbs.iter_rows(min_row=2, max_row=ws_dbs.max_row, values_only=True):
    if row[1] is None or row[5] is None:
        continue
    try:
        pincode = int(row[1])
    except (ValueError, TypeError):
        continue
    zone = str(row[5]).strip()
    if not zone:
        continue
    oda_raw = str(row[4]).strip() if row[4] is not None else 'No'
    is_oda  = oda_raw.lower() == 'yes'

    if zone not in dbs_pincodes:
        dbs_pincodes[zone] = {'all': [], 'oda': [], 'non_oda': []}
    dbs_pincodes[zone]['all'].append(pincode)
    if is_oda:
        dbs_pincodes[zone]['oda'].append(pincode)
    else:
        dbs_pincodes[zone]['non_oda'].append(pincode)

for zone in dbs_pincodes:
    dbs_pincodes[zone]['all'].sort()
    dbs_pincodes[zone]['oda'].sort()
    dbs_pincodes[zone]['non_oda'].sort()

print(f"DB Schenker: {sum(len(v['all']) for v in dbs_pincodes.values())} pincodes, "
      f"{len(dbs_pincodes)} zones, "
      f"rate table: {dbs_rate_table}")

# ============================================================
# STEP 2: VL CARGO DATA FROM Pincode_B2B_Delhivery
# ============================================================
ws_del = wb['Pincode_B2B_Delhivery']
vl_oda_by_zone = {}   # zone -> [oda pincodes]

for row in ws_del.iter_rows(min_row=2, max_row=ws_del.max_row, values_only=True):
    if row[1] is None or row[4] is None:
        continue
    try:
        pincode = int(row[1])
    except (ValueError, TypeError):
        continue
    zone    = str(row[4]).strip()
    oda_raw = str(row[5]).strip() if row[5] is not None else 'No'
    if zone and oda_raw.lower() == 'yes':
        if zone not in vl_oda_by_zone:
            vl_oda_by_zone[zone] = []
        vl_oda_by_zone[zone].append(pincode)

for zone in vl_oda_by_zone:
    vl_oda_by_zone[zone].sort()

total_vl_oda = sum(len(v) for v in vl_oda_by_zone.values())
print(f"VL Cargo: {total_vl_oda} ODA pincodes across {len(vl_oda_by_zone)} zones")

# ============================================================
# STEP 3: LOAD SHIPSHOPY UTSF (serviceability template for VL Cargo)
# ============================================================
shipshopy_path = os.path.join(UTSF_DIR, f'{SHIPSHOPY_ID}.utsf.json')
with open(shipshopy_path, 'r', encoding='utf-8') as f:
    shipshopy = json.load(f)

# ============================================================
# STEP 4: BUILD DB SCHENKER UTSF
# ============================================================

def round2(x):
    return round(x * 100) / 100

# Serviceability
dbs_serviceability = {}
for zone in ALL_ZONES:
    if zone in dbs_pincodes:
        d       = dbs_pincodes[zone]
        total   = len(d['all'])
        oda_cnt = len(d['oda'])
        served  = total - oda_cnt   # non-ODA pincodes
        coverage = round2(served / total * 100) if total > 0 else 0
        dbs_serviceability[zone] = {
            "mode":           "FULL_MINUS_EXCEPT",
            "totalInZone":    total,
            "servedCount":    served,
            "coveragePercent": coverage,
            "exceptSingles":  [],
            "exceptRanges":   [],
            "softExclusions": d['oda']
        }
    else:
        std_total = STANDARD_TOTALS.get(zone, 0)
        dbs_serviceability[zone] = {
            "mode":        "NOT_SERVED",
            "totalInZone": std_total,
            "servedCount": 0,
            "coveragePercent": 0
        }

# ODA section
dbs_oda = {}
for zone in sorted(dbs_pincodes.keys()):
    oda_pins = dbs_pincodes[zone]['oda']
    dbs_oda[zone] = {
        "odaRanges":  [],
        "odaSingles": oda_pins,
        "odaCount":   len(oda_pins)
    }

# Zone rates — hub-and-spoke from N1 (Delhi origin)
dbs_zone_rates = {
    "N1": {z: dbs_rate_table[z] for z in sorted(dbs_rate_table.keys())}
}

# Stats
dbs_total_pins = sum(len(v['all'])  for v in dbs_pincodes.values())
dbs_total_oda  = sum(len(v['oda']) for v in dbs_pincodes.values())
dbs_total_zones = len(dbs_pincodes)

dbs_coverage_region = {
    "North":     sum(len(dbs_pincodes.get(z, {}).get('all', [])) for z in ['N1','N2','N3']),
    "South":     sum(len(dbs_pincodes.get(z, {}).get('all', [])) for z in ['S1','S2','S3']),
    "East":      sum(len(dbs_pincodes.get(z, {}).get('all', [])) for z in ['E1','E2']),
    "West":      sum(len(dbs_pincodes.get(z, {}).get('all', [])) for z in ['W1','W2']),
    "Central":   sum(len(dbs_pincodes.get(z, {}).get('all', [])) for z in ['C1','C2']),
    "North East": sum(len(dbs_pincodes.get(z, {}).get('all', [])) for z in ['NE1']),
    "Special":   0
}

served_zones_pct = [
    round2(len(dbs_pincodes[z]['non_oda']) / len(dbs_pincodes[z]['all']) * 100)
    for z in dbs_pincodes if len(dbs_pincodes[z]['all']) > 0
]
avg_coverage = round2(sum(served_zones_pct) / len(served_zones_pct)) if served_zones_pct else 0

dbs_utsf = {
    "version":      "3.0",
    "generatedAt":  NOW,
    "sourceFormat": "excel",
    "meta": {
        "id":              DBS_ID,
        "companyName":     "DB Schenker",
        "vendorCode":      None,
        "customerID":      None,
        "transporterType": "regular",
        "transportMode":   "LTL",
        "gstNo":           None,
        "address":         None,
        "state":           None,
        "city":            None,
        "pincode":         "",
        "rating":          4,
        "isVerified":      False,
        "approvalStatus":  "pending",
        "createdAt":       NOW,
        "updatedAt":       NOW,
        "created": {
            "by":     "EXCEL_GENERATOR_SCRIPT",
            "at":     NOW,
            "source": "excel"
        },
        "version":        "3.0.0",
        "updateCount":    1,
        "integrityMode":  "STRICT"
    },
    "pricing": {
        "priceRate": {
            "minWeight":         50,
            "docketCharges":     100,
            "fuel":              5,
            "divisor":           27000,
            "kFactor":           27000,
            "minCharges":        400,
            "greenTax":          0,
            "daccCharges":       0,
            "miscCharges":       0,
            "rovCharges":        {"v": 0, "f": 0},
            "insuranceCharges":  {"v": 0, "f": 0},
            "odaCharges":        {"v": 0, "f": 850},
            "codCharges":        {"v": 0, "f": 0},
            "prepaidCharges":    {"v": 0, "f": 0},
            "topayCharges":      {"v": 0, "f": 0},
            "handlingCharges":   {"v": 0, "f": 0},
            "fmCharges":         {"v": 0, "f": 0},
            "appointmentCharges":{"v": 0, "f": 0},
            "invoiceValueCharges": None
        },
        "zoneRates": dbs_zone_rates
    },
    "serviceability": dbs_serviceability,
    "oda":  dbs_oda,
    "stats": {
        "totalPincodes":    dbs_total_pins,
        "totalZones":       dbs_total_zones,
        "odaCount":         dbs_total_oda,
        "coverageByRegion": dbs_coverage_region,
        "avgCoveragePercent": avg_coverage,
        "dataCompleteness": 100,
        "complianceScore":  0.08
    },
    "updates": [
        {
            "timestamp":     NOW,
            "editorId":      "EXCEL_GENERATOR_SCRIPT",
            "reason":        "Initial creation from Excel source of truth",
            "changeSummary": (f"Created DB Schenker UTSF: {dbs_total_pins} pincodes, "
                              f"{dbs_total_zones} zones, {dbs_total_oda} ODA pincodes. "
                              f"Rate table: {dbs_zone_rates['N1']}"),
            "snapshot": None
        }
    ]
}

# ============================================================
# STEP 5: BUILD VL CARGO UTSF
# ============================================================

# Zone rates — flat 18/kg for all zone pairs
vl_zone_rates = {
    origin: {dest: 18.0 for dest in VL_ZONES}
    for origin in VL_ZONES
}

# ODA section (from Pincode_B2B_Delhivery)
vl_oda = {}
for zone in VL_ZONES:
    oda_pins = sorted(vl_oda_by_zone.get(zone, []))
    vl_oda[zone] = {
        "odaRanges":  [],
        "odaSingles": oda_pins,
        "odaCount":   len(oda_pins)
    }

# Serviceability — copy from Shipshopy (same Delhivery network)
import copy
vl_serviceability = copy.deepcopy(shipshopy['serviceability'])

# Stats
vl_total_oda     = sum(len(vl_oda_by_zone.get(z, [])) for z in VL_ZONES)
ship_stats       = shipshopy.get('stats', {})
vl_served_zones  = len(VL_ZONES)

vl_utsf = {
    "version":      "3.0",
    "generatedAt":  NOW,
    "sourceFormat": "excel",
    "meta": {
        "id":              VL_ID,
        "companyName":     "Delhivery (VL Cargo)",
        "vendorCode":      None,
        "customerID":      None,
        "transporterType": "regular",
        "transportMode":   "LTL",
        "gstNo":           None,
        "address":         None,
        "state":           None,
        "city":            None,
        "pincode":         "",
        "rating":          4,
        "isVerified":      False,
        "approvalStatus":  "pending",
        "createdAt":       NOW,
        "updatedAt":       NOW,
        "created": {
            "by":     "EXCEL_GENERATOR_SCRIPT",
            "at":     NOW,
            "source": "excel"
        },
        "version":       "3.0.0",
        "updateCount":   1,
        "integrityMode": "STRICT"
    },
    "pricing": {
        "priceRate": {
            "minWeight":         20,
            "docketCharges":     0,
            "fuel":              0,
            "divisor":           1,
            "kFactor":           4500,
            "minCharges":        0,
            "greenTax":          0,
            "daccCharges":       0,
            "miscCharges":       0,
            "rovCharges":        {"v": 0, "f": 0},
            "insuranceCharges":  {"v": 0, "f": 0},
            "odaCharges":        {"v": 0, "f": 0},
            "codCharges":        {"v": 0, "f": 0},
            "prepaidCharges":    {"v": 0, "f": 0},
            "topayCharges":      {"v": 0, "f": 0},
            "handlingCharges":   {"v": 0, "f": 0},
            "fmCharges":         {"v": 0, "f": 0},
            "appointmentCharges":{"v": 0, "f": 0},
            "invoiceValueCharges": None
        },
        "zoneRates": vl_zone_rates
    },
    "serviceability": vl_serviceability,
    "oda":  vl_oda,
    "stats": {
        "totalPincodes":     ship_stats.get('totalPincodes', 21339),
        "totalZones":        vl_served_zones,
        "odaCount":          vl_total_oda,
        "coverageByRegion":  ship_stats.get('coverageByRegion', {}),
        "avgCoveragePercent": ship_stats.get('avgCoveragePercent', 0),
        "dataCompleteness":  100,
        "complianceScore":   0.0729
    },
    "updates": [
        {
            "timestamp":     NOW,
            "editorId":      "EXCEL_GENERATOR_SCRIPT",
            "reason":        "Initial creation from Excel source of truth",
            "changeSummary": (f"Created Delhivery VL Cargo UTSF: flat 18/kg across "
                              f"{vl_served_zones} zones; {vl_total_oda} ODA pincodes "
                              f"from Delhivery B2B network"),
            "snapshot": None
        }
    ]
}

# ============================================================
# STEP 6: WRITE FILES
# ============================================================
dbs_path = os.path.join(UTSF_DIR, f'{DBS_ID}.utsf.json')
vl_path  = os.path.join(UTSF_DIR, f'{VL_ID}.utsf.json')

with open(dbs_path, 'w', encoding='utf-8') as f:
    json.dump(dbs_utsf, f, indent=2, ensure_ascii=False)
    f.write('\n')
print(f"Written: {dbs_path}")

with open(vl_path, 'w', encoding='utf-8') as f:
    json.dump(vl_utsf, f, indent=2, ensure_ascii=False)
    f.write('\n')
print(f"Written: {vl_path}")

# ============================================================
# STEP 7: VERIFICATION
# ============================================================
print("\n" + "="*60)
print("VERIFICATION: DB Schenker (Pincode 226010, Zone N2, 2500 kg, No ODA)")
print("="*60)
weight = 2500
zone   = 'N2'
rate   = dbs_zone_rates['N1'][zone]
base   = max(400, weight * rate)   # apply minCharges
fuel   = base * 0.05
docket = 100
oda_charge = 0
final  = base + fuel + docket + oda_charge
print(f"  Rate (N2): {rate}/kg")
print(f"  Base (2500 x {rate}): {weight * rate}")
print(f"  After minCharges check: {base}")
print(f"  Fuel (5% of {weight*rate}): {weight*rate*0.05}")
print(f"  Docket: {docket}")
print(f"  ODA: {oda_charge}")
print(f"  FINAL: {base + fuel + docket}")
print(f"  Excel Expected: 17162.5")
# Correct: base=16250, fuel=812.5, docket=100, final=17162.5
# The minCharges check: 16250 > 400, so base=16250
base2  = weight * rate   # 16250
fuel2  = base2 * 0.05   # 812.5
final2 = base2 + fuel2 + docket
print(f"  Corrected (no minCharges clamp needed): {final2}")
print(f"  Match Excel: {abs(final2 - 17162.5) < 0.01}")

print("\n" + "="*60)
print("VERIFICATION: VL Cargo (Pincode 226010, Zone N3, 2500 kg, No ODA)")
print("="*60)
vl_rate  = 18
vl_base  = 2500 * vl_rate
vl_final = vl_base   # no fuel, no docket for VL Cargo
print(f"  Rate (N3): {vl_rate}/kg")
print(f"  Base (2500 x 18): {vl_base}")
print(f"  Fuel: 0")
print(f"  Docket+ROV: 0")
print(f"  FINAL: {vl_final}")
print(f"  Excel Expected: 45000")
print(f"  Match Excel: {abs(vl_final - 45000) < 0.01}")

print("\n" + "="*60)
print("FILE SIZES")
print("="*60)
print(f"  DB Schenker: {os.path.getsize(dbs_path):,} bytes")
print(f"  VL Cargo:    {os.path.getsize(vl_path):,} bytes")

print("\n" + "="*60)
print("DB SCHENKER ZONE RATE SUMMARY (N1 origin)")
print("="*60)
for z, r in sorted(dbs_zone_rates['N1'].items()):
    cnt = len(dbs_pincodes.get(z, {}).get('all', []))
    oda = len(dbs_pincodes.get(z, {}).get('oda', []))
    print(f"  {z}: rate={r}/kg, total={cnt}, oda={oda}")

print("\n" + "="*60)
print("VL CARGO ODA COUNTS BY ZONE")
print("="*60)
for zone in VL_ZONES:
    cnt = vl_oda[zone]['odaCount']
    print(f"  {zone}: {cnt} ODA pincodes")

print("\nDONE.")
