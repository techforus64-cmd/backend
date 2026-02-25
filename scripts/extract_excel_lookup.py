"""
extract_excel_lookup.py
Reads Pincode_B2B_Delhivery and Pincode_DBS sheets from the Excel file
and saves them as JSON for the comprehensive UTSF test.
"""
import openpyxl
import json
import os

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'Transport Cost Calculator (5).xlsx')
OUT_PATH   = os.path.join(os.path.dirname(__file__), 'excel_lookup.json')

print("Reading Excel file ...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)

# ── Pincode_B2B_Delhivery ─────────────────────────────────────────────────
# Cols (0-indexed from row tuple):
#   0=S No, 1=PINCODE, 2=Facility City, 3=Facility State,
#   4=Zone, 5=ODA, 6=SS_price, 7=VLC_price, 8=DBS_price(unused), 9=SFX_price
ws1 = wb['Pincode_B2B_Delhivery']
delhivery = {}
skipped = 0
for row in ws1.iter_rows(min_row=2, values_only=True):
    pin_raw = row[1]
    if pin_raw is None:
        continue
    try:
        pin = int(pin_raw)
    except (ValueError, TypeError):
        skipped += 1
        continue
    zone    = row[4]
    oda     = str(row[5] or 'No').strip().upper()   # 'Yes' or 'No'
    ss_p    = row[6]   # Shipshopy unit price
    vlc_p   = row[7]   # VL Cargo unit price
    sfx_p   = row[9]   # Safexpress unit price  (col J = index 9)
    if zone and ss_p is not None:
        delhivery[pin] = {
            'zone': str(zone).strip().upper(),
            'oda':  'Yes' if oda.startswith('Y') else 'No',
            'ss':   float(ss_p),
            'vlc':  float(vlc_p) if vlc_p is not None else None,
            'sfx':  float(sfx_p) if sfx_p is not None else None,
        }

print(f"  Delhivery lookup: {len(delhivery)} pincodes (skipped {skipped})")

# ── Pincode_DBS ───────────────────────────────────────────────────────────
# Cols: 0=S No, 1=Pincode, 2=City, 3=State, 4=ODA, 5=Zone
# Zone rate table: I(col 8) = zone name, J(col 9) = price
ws2 = wb['Pincode_DBS']

dbs_rates = {}
dbs_pincodes = {}
for row in ws2.iter_rows(min_row=2, values_only=True):
    # Zone rate table in cols I-J (index 8-9)
    if row[8] is not None and row[9] is not None:
        dbs_rates[str(row[8]).strip().upper()] = float(row[9])

    pin_raw = row[1]
    if pin_raw is None:
        continue
    try:
        pin = int(pin_raw)
    except (ValueError, TypeError):
        continue
    oda  = str(row[4] or 'No').strip().upper()
    zone = row[5]
    if zone:
        dbs_pincodes[pin] = {
            'oda':  'Yes' if oda.startswith('Y') else 'No',
            'zone': str(zone).strip().upper(),
        }

print(f"  DBS lookup: {len(dbs_pincodes)} pincodes")
print(f"  DBS zone rates: {dbs_rates}")

wb.close()

out = {
    'delhivery': delhivery,   # keys are int pincodes as strings (JSON)
    'dbs':       dbs_pincodes,
    'dbs_rates': dbs_rates,
}

# JSON keys must be strings
out_str_keys = {
    'delhivery': {str(k): v for k, v in delhivery.items()},
    'dbs':       {str(k): v for k, v in dbs_pincodes.items()},
    'dbs_rates': dbs_rates,
}
with open(OUT_PATH, 'w') as f:
    json.dump(out_str_keys, f, indent=2)

print(f"Saved -> {OUT_PATH}")
